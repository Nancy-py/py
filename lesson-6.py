'''
python自动化工作：
1、准备好自动化测试用例  ==== done
2、使用python去读取测试用例
3、发送请求，得到响应结果 ====done
4、结果的判断，执行结果  vs  预期结果====断言
5、得到一个最终结果，回写到测试用例
'''
# 如果我们要用python去执行一个功能，先去搜索一下，有没有对于的第三方库支持
# 读取Excel的第三方库，openpyxl
# 先安装，在导入
# pip安装   python安装
'''
 第三方库:
 1、requests  ==  发送http请求得到响应结果
 2、jsonpath  ==  可以做关联，取出接口响应结果里的数据
 3、openpyxl  ==  读取Excel并回写数据
'''
# import openpyxl
# # python操作Excel的三大对象：
# # 1、加载工作溥（Excel）
# wb = openpyxl.load_workbook('test_case_api.xlsx') #==加载工作溥到py的内存中
# print(wb)
# # 2、操作sheet（Excel中的）表单
# sheet = wb['register']
# print(sheet )
# # 3、操作单元格
# # row =行  column = 列、
# cell = sheet.cell(row=1,column=1) #读取cell
# print(cell) #这样取出的是单元格
# # 取出单元格的数据，要加上.value
# print(cell.value)
#
# # 写入数据
# # value赋值  （修改）
# cell.value = '测试用例编号'
# print(cell.value)
# # 把写入的数据进行保存
# # 在保存之前一定要关闭Excel，不然无法保存或报错
# wb.save('test_case_api.xlsx')

# 自动读取excel
# import requests
# import openpyxl
# wb = openpyxl.load_workbook('test_case_api.xlsx')
# sheet = wb['login']
# url = sheet.cell(row=2,column=5).value #读取url==上面读取cell一样
# data = sheet.cell(row=2,column=6).value #读请求体
# expected = sheet.cell(row=2,column=7).value #读取预期结果
# print(url,data,expected)
#
# def login_func(url,data):
#   header_login = {'X-Lemonban-Media-Type':'lemonban.v2',
#   'Content-Type':'application/json'}  #请求头
#   res1 = requests.post(url=url,json=data,headers=header_login)
#   print(res1.json())
#
# res = login_func(url,data)
# print(res)

# import requests
# import openpyxl
# wb = openpyxl.load_workbook('test_case_api.xlsx')
# sheet = wb['login']
# for i in range(2,8,1):
#     url = sheet.cell(row=2,column=5).value #读取url==上面读取cell一样
#     data = sheet.cell(row=2,column=6).value #读请求体
#     expected = sheet.cell(row=2,column=7).value #读取预期结果
#     print(url,data,expected)

# max_row  取最大的行数
# max_column  取最大列数
# import requests
# import openpyxl
# wb = openpyxl.load_workbook('test_case_api.xlsx')
# sheet = wb['login']
# max_row = sheet.max_row #取出sheet里面最大的行数
# # print(max_row)
# for i in range(2,max_row+1,1): #取左不取右，左闭右开 所以+1
#     url = sheet.cell(row=2,column=5).value #读取url==上面读取cell一样
#     data = sheet.cell(row=2,column=6).value #读请求体
#     expected = sheet.cell(row=2,column=7).value #读取预期结果
#     print(url,data,expected)

# import requests
# import openpyxl
# wb = openpyxl.load_workbook('test_case_api.xlsx')
# sheet = wb['login']
# max_row = sheet.max_row #取出sheet里面最大的行数
# # print(max_row)
# for i in range(2,max_row+1,1):
#     dict1 = dict(
#     url = sheet.cell(row=2,column=5).value,
#     data = sheet.cell(row=2,column=6).value ,
#     expected = sheet.cell(row=2,column=7).value)
#     print(dict1)   #字典格式

import requests
import openpyxl
# wb = openpyxl.load_workbook('test_case_api.xlsx')
# sheet = wb['login']
# max_row = sheet.max_row #取出sheet里面最大的行数
# # print(max_row)
# cose_list = []  #定义一个空列表
# for i in range(2,max_row+1,1):
#     dict1 = dict(
#     url = sheet.cell(row=2,column=5).value,
#     data = sheet.cell(row=2,column=6).value ,
#     expected = sheet.cell(row=2,column=7).value)
#     cose_list.append(dict1)#把dict1里面一条一条的测试用例装到列表里面，这个列表里面存放了所有测试用例
# print(cose_list)

# import requests
# import openpyxl
# def read_data(filename,sheetname):
#     wb = openpyxl.load_workbook(filename)  #改名为
#     sheet = wb[sheetname]
#     max_row = sheet.max_row #取出sheet里面最大的行数
#     # print(max_row)
#     case_list = []
#     for i in range(2,max_row+1,1):
#         dict1 = dict(
#         case_id=sheet.cell(row=i,column=1).value,
#         url = sheet.cell(row=i,column=5).value,
#         data = sheet.cell(row=i,column=6).value ,
#         expected = sheet.cell(row=i,column=7).value)
#         case_list.append(dict1)#把dict1里面一条一条的测试用例装到列表里面，这个列表里面存放了所有测试用例
#     # print(case_list)
#     return case_list  #一定要return  返回值
# case = read_data('test_case_api.xlsx','register')
# print(case)


wb = openpyxl.load_workbook('test_case_api.xlsx')
sheet = wb['register']
sheet.cell(row=2,column=8).value = 'pass'
wb.save('test_case_api.xlsx') #把pass保存到Excel中

# 把有变化的参数封装成函数
def write_result(filename,sheetname,row,column,final_result):
    wb = openpyxl.load_workbook(filename)
    sheet = wb[sheetname]
    sheet.cell(row=row,column=column).value = final_result
    wb.save(filename)
case = write_result('test_case_api.xlsx','register',row =3,column=8,final_result='pass')
print(case)